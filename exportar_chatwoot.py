import requests
import json
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# ====== CONFIGURA ESTO ======
CHATWOOT_URL = "https://chatwoot.educarteonline.lat"  # Cambia si es diferente
API_TOKEN = "Ty8626Djdsm74aYcH5gpc9pb"                       # Tu access token
ACCOUNT_ID = 1
# ============================

headers = {"api_access_token": API_TOKEN}
all_contacts = []
page = 1

print("Descargando contactos...")
while True:
    url = f"{CHATWOOT_URL}/api/v1/accounts/{ACCOUNT_ID}/contacts?page={page}&sort=created_at"
    r = requests.get(url, headers=headers)
    data = r.json()
    contacts = data.get("payload", [])
    if not contacts:
        break
    all_contacts.extend(contacts)
    total = data.get("meta", {}).get("count", "?")
    print(f"  Página {page}: {len(all_contacts)}/{total} contactos")
    page += 1

print(f"\nTotal descargados: {len(all_contacts)}")

# Obtener labels de cada contacto
print("Descargando etiquetas de cada contacto...")
for i, contact in enumerate(all_contacts):
    cid = contact["id"]
    try:
        url = f"{CHATWOOT_URL}/api/v1/accounts/{ACCOUNT_ID}/contacts/{cid}/labels"
        r = requests.get(url, headers=headers)
        labels = r.json().get("payload", [])
        contact["_labels"] = ", ".join(labels) if labels else ""
    except:
        contact["_labels"] = ""
    if (i+1) % 20 == 0:
        print(f"  {i+1}/{len(all_contacts)}")

print(f"  {len(all_contacts)}/{len(all_contacts)} - Listo")

# Crear Excel
wb = Workbook()
ws = wb.active
ws.title = "Contactos Chatwoot"

# Headers
columnas = [
    "Nombre", "Teléfono", "Email", "Labels",
    "País", "Producto", "Producto ID", "Etapa",
    "Ad Source ID", "Ad Headline", "Workflow",
    "Fecha Creación", "Última Actividad",
    "Inbox"
]

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col, name in enumerate(columnas, 1):
    cell = ws.cell(row=1, column=col, value=name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Data
for i, c in enumerate(all_contacts, 2):
    ca = c.get("custom_attributes", {}) or {}
    inbox_name = ""
    inboxes = c.get("contact_inboxes", [])
    if inboxes:
        inbox_name = inboxes[0].get("inbox", {}).get("name", "")
    
    created = ""
    if c.get("created_at"):
        try:
            created = datetime.fromtimestamp(c["created_at"]).strftime("%Y-%m-%d %H:%M")
        except:
            created = str(c["created_at"])
    
    last_activity = ""
    if c.get("last_activity_at"):
        try:
            last_activity = datetime.fromtimestamp(c["last_activity_at"]).strftime("%Y-%m-%d %H:%M")
        except:
            last_activity = str(c["last_activity_at"])

    ws.cell(row=i, column=1, value=c.get("name", ""))
    ws.cell(row=i, column=2, value=c.get("phone_number", ""))
    ws.cell(row=i, column=3, value=c.get("email", ""))
    ws.cell(row=i, column=4, value=c.get("_labels", ""))
    ws.cell(row=i, column=5, value=ca.get("pais", ""))
    ws.cell(row=i, column=6, value=ca.get("producto", ""))
    ws.cell(row=i, column=7, value=ca.get("producto_id", ""))
    ws.cell(row=i, column=8, value=ca.get("etapa", ""))
    ws.cell(row=i, column=9, value=ca.get("ad_source_id", ""))
    ws.cell(row=i, column=10, value=ca.get("ad_headline", ""))
    ws.cell(row=i, column=11, value=ca.get("workflow", ""))
    ws.cell(row=i, column=12, value=created)
    ws.cell(row=i, column=13, value=last_activity)
    ws.cell(row=i, column=14, value=inbox_name)

# Ajustar anchos
for col in range(1, len(columnas) + 1):
    ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 18

# Filtros
ws.auto_filter.ref = ws.dimensions

filename = f"chatwoot_contactos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(filename)
print(f"\n✅ Excel guardado: {filename}")
print(f"   Ruta completa: {filename}")
print(f"   Abre con: open '{filename}'")
